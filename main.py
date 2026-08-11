from __future__ import annotations

import asyncio
import csv
from io import TextIOWrapper
from pathlib import Path
from typing import Any

from astrbot.api import logger
from astrbot.api.event import AstrMessageEvent, filter
from astrbot.api.star import Context, Star, register
from astrbot.api.web import PluginUploadFile, error_response, json_response, request
from astrbot.core.utils.astrbot_path import get_astrbot_plugin_data_path

from .game import DECK_SIZE, GameError, ParcelGame, initial_state

PLUGIN_NAME = "astrbot_plugin_zombie_station"
STATE_KEY = "zombie_station_state"


@register(PLUGIN_NAME, "local", "QQ 群末日快递驿站开包副本", "0.1.0")
class ZombieStationPlugin(Star):
    def __init__(self, context: Context, config: dict[str, Any]):
        super().__init__(context)
        self.config = config
        self.lock = asyncio.Lock()
        self.game = ParcelGame(self._daily_stamina())
        context.register_web_api(f"/{PLUGIN_NAME}/stats", self.web_stats, ["GET"], "驿站统计")
        context.register_web_api(f"/{PLUGIN_NAME}/cards/import", self.web_import_cards, ["POST"], "导入卡牌表")

    def _daily_stamina(self) -> int:
        return max(1, int(self.config.get("daily_stamina", 10)))

    async def _load_state(self) -> dict[str, Any]:
        state = await self.get_kv_data(STATE_KEY, None)
        if not isinstance(state, dict) or not isinstance(state.get("cards"), list):
            return initial_state()
        if len(state["cards"]) != DECK_SIZE:
            logger.warning("驿站卡池数量不是 1200，已恢复内置占位卡池。")
            return initial_state()
        state.setdefault("groups", {})
        return state

    async def _save_state(self, state: dict[str, Any]) -> None:
        await self.put_kv_data(STATE_KEY, state)

    @staticmethod
    def _group_id(event: AstrMessageEvent) -> str:
        return str(event.get_group_id() or event.get_session_id())

    @staticmethod
    def _sender_name(event: AstrMessageEvent) -> str:
        return event.get_sender_name() or event.get_sender_id()

    @filter.command_group("驿站")
    def station(self):
        pass

    @station.command("开局")
    @filter.event_message_type(filter.EventMessageType.GROUP_MESSAGE)
    async def start_game(self, event: AstrMessageEvent):
        async with self.lock:
            state = await self._load_state()
            try:
                group = self.game.start(state, self._group_id(event), event.get_sender_id(), self._sender_name(event))
            except GameError as exc:
                yield event.plain_result(f"【快递驿站】{exc}")
                return
            await self._save_state(state)
        yield event.plain_result(
            f"【快递驿站】第 {group['round']} 局开始。{group['host_name']} 是主持人。\n"
            f"每人每天 {self._daily_stamina()} 点体力；发送“驿站 开包 [1-10]”开包，省略数量默认开 10 个。"
        )

    @station.command("开包")
    @filter.event_message_type(filter.EventMessageType.GROUP_MESSAGE)
    async def draw_parcels(self, event: AstrMessageEvent, amount: int = 10):
        async with self.lock:
            state = await self._load_state()
            self.game = ParcelGame(self._daily_stamina())
            try:
                result = self.game.draw(state, self._group_id(event), event.get_sender_id(), self._sender_name(event), amount)
            except GameError as exc:
                yield event.plain_result(f"【快递驿站】{exc}")
                return
            await self._save_state(state)
        lines = [f"【快递驿站｜第 {result.round_number} 局｜第 {result.day} 天】"]
        lines.extend(f"{index}. {card['title']}\n{card['text']}" for index, card in enumerate(result.cards, start=1))
        lines.append(f"体力剩余：{result.stamina_left}｜本局包裹剩余：{result.remaining}/{DECK_SIZE}")
        if result.completed_round:
            lines.append(f"第 {result.completed_round} 局的 1200 个包裹已全部抽光；第 {result.completed_round + 1} 局已自动重置。")
        yield event.plain_result("\n".join(lines))

    @station.command("新一天")
    @filter.event_message_type(filter.EventMessageType.GROUP_MESSAGE)
    async def advance_day(self, event: AstrMessageEvent):
        async with self.lock:
            state = await self._load_state()
            self.game = ParcelGame(self._daily_stamina())
            try:
                group = self.game.next_day(state, self._group_id(event), event.get_sender_id())
            except GameError as exc:
                yield event.plain_result(f"【快递驿站】{exc}")
                return
            await self._save_state(state)
        yield event.plain_result(f"【快递驿站】第 {group['round']} 局进入第 {group['day']} 天；所有已入局玩家体力均恢复到各自的每日上限。")

    @station.command("设体力")
    @filter.event_message_type(filter.EventMessageType.GROUP_MESSAGE)
    async def set_player_stamina(self, event: AstrMessageEvent, player_id: str, daily_stamina: int):
        async with self.lock:
            state = await self._load_state()
            self.game = ParcelGame(self._daily_stamina())
            try:
                player = self.game.set_player_stamina(
                    state, self._group_id(event), event.get_sender_id(), player_id, daily_stamina
                )
            except GameError as exc:
                yield event.plain_result(f"【快递驿站】{exc}")
                return
            await self._save_state(state)
        yield event.plain_result(
            f"【快递驿站】已将玩家 {player_id} 的每日体力设为 {player['daily_stamina']}；当前体力已同步为该数值。"
        )

    @station.command("状态")
    @filter.event_message_type(filter.EventMessageType.GROUP_MESSAGE)
    async def game_status(self, event: AstrMessageEvent):
        async with self.lock:
            state = await self._load_state()
            try:
                group = self.game.get_group(state, self._group_id(event))
            except GameError as exc:
                yield event.plain_result(f"【快递驿站】{exc}")
                return
            player = group["players"].get(event.get_sender_id())
        personal = "你尚未开过包。" if player is None else (
            f"你的体力：{player['stamina']}/{player.get('daily_stamina', self._daily_stamina())}｜"
            f"今日已开：{player['opened_today']}｜累计：{player['total_opened']}"
        )
        yield event.plain_result(f"【快递驿站】第 {group['round']} 局，第 {group['day']} 天\n主持人：{group['host_name']}｜包裹剩余：{len(group['deck'])}/{DECK_SIZE}\n{personal}")

    @station.command("帮助")
    @filter.event_message_type(filter.EventMessageType.GROUP_MESSAGE)
    async def game_help(self, event: AstrMessageEvent):
        yield event.plain_result(
            "【快递驿站指令】\n"
            "驿站 开局：创建副本，发起者成为主持人。\n"
            "驿站 开包 [1-10]：消耗等量体力开包；省略数量时开 10 个。\n"
            "驿站 新一天：仅主持人可用，按每名玩家的每日上限恢复体力。\n"
            "驿站 设体力 <QQ号> <数值>：仅主持人可用，设置该玩家每日体力。\n"
            "驿站 状态：查看本局进度与自己的体力。\n"
            "驿站 管理重置：AstrBot 管理员强制开下一局。\n"
            "驿站 管理设体力 <QQ号> <数值>：AstrBot 管理员强制设置玩家每日体力。\n"
            "驿站 管理状态：AstrBot 管理员查看本群汇总。"
        )

    @station.command("管理重置")
    @filter.permission_type(filter.PermissionType.ADMIN)
    @filter.event_message_type(filter.EventMessageType.GROUP_MESSAGE)
    async def admin_reset(self, event: AstrMessageEvent):
        async with self.lock:
            state = await self._load_state()
            self.game = ParcelGame(self._daily_stamina())
            group = self.game.reset(state, self._group_id(event), event.get_sender_id(), self._sender_name(event))
            await self._save_state(state)
        yield event.plain_result(f"【快递驿站】管理员已重置为第 {group['round']} 局，第 1 天。")

    @station.command("管理设体力")
    @filter.permission_type(filter.PermissionType.ADMIN)
    @filter.event_message_type(filter.EventMessageType.GROUP_MESSAGE)
    async def admin_set_player_stamina(self, event: AstrMessageEvent, player_id: str, daily_stamina: int):
        async with self.lock:
            state = await self._load_state()
            self.game = ParcelGame(self._daily_stamina())
            try:
                player = self.game.set_player_stamina(
                    state, self._group_id(event), event.get_sender_id(), player_id, daily_stamina, force=True
                )
            except GameError as exc:
                yield event.plain_result(f"【快递驿站】{exc}")
                return
            await self._save_state(state)
        yield event.plain_result(
            f"【快递驿站】管理员已将玩家 {player_id} 的每日体力设为 {player['daily_stamina']}。"
        )

    @station.command("管理状态")
    @filter.permission_type(filter.PermissionType.ADMIN)
    @filter.event_message_type(filter.EventMessageType.GROUP_MESSAGE)
    async def admin_status(self, event: AstrMessageEvent):
        async with self.lock:
            state = await self._load_state()
            snapshot = self.game.snapshot(state, self._group_id(event))
        if not snapshot["groups"]:
            yield event.plain_result("【快递驿站】本群尚未开局。")
            return
        group = snapshot["groups"][0]
        rankings = "、".join(
            f"{player['name']} {player['total_opened']}包/日上限{player.get('daily_stamina', self._daily_stamina())}"
            for player in group["players"][:5]
        ) or "暂无玩家"
        yield event.plain_result(f"【快递驿站管理】卡池：{snapshot['cards_source']}\n第 {group['round']} 局｜第 {group['day']} 天｜剩余 {group['remaining']}/{DECK_SIZE}\n玩家 {group['player_count']} 人｜累计开包排行：{rankings}")

    async def web_stats(self):
        async with self.lock:
            state = await self._load_state()
            snapshot = self.game.snapshot(state)
        return json_response(snapshot)

    async def web_import_cards(self):
        files = await request.files()
        upload = files.get("file")
        if not isinstance(upload, PluginUploadFile):
            return error_response("请选择一个 CSV 或 XLSX 文件。", status_code=400)
        filename = upload.filename or ""
        suffix = Path(filename).suffix.lower()
        if suffix not in {".csv", ".xlsx"}:
            return error_response("仅支持 .csv 和 .xlsx 文件。", status_code=400)
        if upload.content_length and upload.content_length > 5 * 1024 * 1024:
            return error_response("文件不能超过 5 MB。", status_code=400)

        import_dir = Path(get_astrbot_plugin_data_path()) / PLUGIN_NAME / "imports"
        import_dir.mkdir(parents=True, exist_ok=True)
        target = import_dir / f"cards{suffix}"
        await upload.save(target)
        try:
            cards = self._read_cards(target)
        except (OSError, UnicodeError, ValueError) as exc:
            return error_response(f"无法读取卡牌表：{exc}", status_code=400)
        if len(cards) != DECK_SIZE:
            return error_response(f"卡牌表必须恰好包含 {DECK_SIZE} 张有效卡牌，当前为 {len(cards)} 张。", status_code=400)

        async with self.lock:
            state = await self._load_state()
            state["cards"] = cards
            state["cards_source"] = Path(filename).name
            state["groups"] = {}
            await self._save_state(state)
        return json_response({"imported": len(cards), "source": Path(filename).name, "cleared_groups": True})

    @staticmethod
    def _read_cards(path: Path) -> list[dict[str, str]]:
        if path.suffix.lower() == ".csv":
            with path.open("rb") as raw_file:
                rows = list(csv.DictReader(TextIOWrapper(raw_file, encoding="utf-8-sig")))
        else:
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=True)
            worksheet = workbook.active
            values = list(worksheet.iter_rows(values_only=True))
            workbook.close()
            if not values:
                raise ValueError("表格为空")
            headers = [str(value or "").strip() for value in values[0]]
            rows = [dict(zip(headers, row, strict=False)) for row in values[1:]]

        cards = []
        for row_number, row in enumerate(rows, start=2):
            normalized = {str(key).strip().lower(): str(value or "").strip() for key, value in row.items()}
            title = next((normalized.get(key, "") for key in ("名称", "标题", "卡牌", "name", "title") if normalized.get(key)), "")
            text = next((normalized.get(key, "") for key in ("内容", "描述", "效果", "text", "description") if normalized.get(key)), "")
            if not title and not text:
                continue
            cards.append({"id": f"P-{len(cards) + 1:04d}", "title": title or f"包裹 {row_number - 1:04d}", "text": text})
        return cards


